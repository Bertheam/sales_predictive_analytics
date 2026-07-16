import hashlib
import io
import re
import unicodedata
from datetime import date, time

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.imports.definitions import IMPORT_DEFINITIONS
from app.repositories.import_repository import ImportRepository


COLUMN_ALIASES = {
    "numero_vente": "sale_reference",
    "sale_number": "sale_reference",
    "reference_vente": "sale_reference",
    "date_vente": "sale_date",
    "heure_vente": "sale_time",
    "code_client": "customer_code",
    "code_produit": "product_code",
    "quantite_colis": "quantity_packages",
    "prix_unitaire": "unit_price",
    "remise": "discount_amount",
    "mode_paiement": "payment_method",
    "statut_paiement": "payment_status",
    "vendeur": "salesperson_name",
    "date_stock": "stock_date",
    "stock_ouverture": "opening_stock",
    "quantite_recue": "quantity_received",
    "quantite_vendue": "quantity_sold",
    "quantite_endommagee": "quantity_damaged",
    "autres_entrees": "other_entries",
    "autres_sorties": "other_outputs",
    "stock_cloture": "closing_stock",
    "nom": "name",
    "marque": "brand",
    "code_categorie": "category_code",
    "type_conditionnement": "package_type",
    "unites_par_colis": "units_per_package",
    "prix_achat": "purchase_price",
    "prix_vente": "selling_price",
    "quantite_reapprovisionnement": "reorder_quantity",
    "code_type_client": "customer_type_code",
    "telephone": "phone",
    "quartier": "district",
    "ville": "city",
}

EXCEL_LISTS = {
    "SALES": {
        "product_code": "products",
        "customer_code": "customers",
        "payment_method": ["CASH", "MOBILE_MONEY", "BANK_TRANSFER", "CREDIT"],
        "payment_status": ["PAID", "PENDING", "PARTIALLY_PAID", "CANCELLED"],
    },
    "STOCKS": {
        "product_code": "products",
    },
    "PRODUCTS": {
        "category_code": "categories",
        "package_type": ["CARTON", "PACK", "CASIER", "UNITE"],
        "volume_unit": ["ML", "CL", "L"],
    },
    "CUSTOMERS": {
        "customer_type_code": "customer_types",
        "city": ["Bamako", "Kati", "Koulikoro"],
    },
}

DATE_COLUMNS = {"sale_date", "stock_date"}
TIME_COLUMNS = {"sale_time"}
INTEGER_COLUMNS = {"units_per_package"}
POSITIVE_COLUMNS = {
    "quantity_packages",
    "unit_price",
    "units_per_package",
    "selling_price",
}
NON_NEGATIVE_COLUMNS = {
    "discount_amount",
    "opening_stock",
    "quantity_received",
    "quantity_sold",
    "quantity_damaged",
    "other_entries",
    "other_outputs",
    "closing_stock",
    "minimum_stock",
    "volume_value",
    "purchase_price",
    "reorder_quantity",
}


class DataImportService:
    def __init__(self, db: Session):
        self.db = db
        self.repository = ImportRepository(db)

    def get_definitions(self) -> dict:
        return IMPORT_DEFINITIONS

    def get_template(
        self,
        import_type: str,
        file_type: str | None = None,
    ) -> bytes:
        if file_type is not None and file_type.upper() != "XLSX":
            raise ValueError("Seul le format Excel XLSX est pris en charge.")
        definition = IMPORT_DEFINITIONS[import_type]
        columns = definition["required"] + definition["optional"]
        example = definition["example"][0]
        references = self.repository.get_reference_maps()

        workbook = Workbook()
        data_sheet = workbook.active
        data_sheet.title = "Données"
        instructions_sheet = workbook.create_sheet("Instructions", 0)
        references_sheet = workbook.create_sheet("Références")

        self._build_instructions_sheet(
            instructions_sheet,
            import_type,
            definition,
        )
        self._build_references_sheet(
            workbook,
            references_sheet,
            import_type,
            references,
        )
        self._build_data_sheet(
            data_sheet,
            import_type,
            definition,
            columns,
            example,
        )
        references_sheet.sheet_state = "hidden"
        workbook.active = 1

        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()

    def analyze_file(
        self,
        *,
        file_name: str,
        content: bytes,
        import_type: str,
    ) -> dict:
        definition = IMPORT_DEFINITIONS[import_type]
        data = self._read_file(file_name, content)
        if data.empty:
            raise ValueError("Le fichier ne contient aucune ligne de données.")

        data.columns = [self._normalize_column(column) for column in data.columns]
        data = data.rename(columns=COLUMN_ALIASES)
        duplicated_columns = data.columns[data.columns.duplicated()].tolist()
        if duplicated_columns:
            raise ValueError(
                "Colonnes présentes plusieurs fois après normalisation : "
                + ", ".join(duplicated_columns)
            )
        missing = [
            column
            for column in definition["required"]
            if column not in data.columns
        ]
        if missing:
            raise ValueError(
                "Colonnes obligatoires manquantes : " + ", ".join(missing)
            )

        allowed = definition["required"] + definition["optional"]
        data = data[[column for column in data.columns if column in allowed]]
        records = []
        for index, raw_row in data.iterrows():
            record = {
                key: self._clean_value(value)
                for key, value in raw_row.to_dict().items()
            }
            record["_row_number"] = index + 2
            record["_errors"] = []
            record["_duplicate"] = False
            records.append(record)

        references = self.repository.get_reference_maps()
        existing = self.repository.get_existing_keys(import_type)
        validators = {
            "SALES": self._validate_sales,
            "STOCKS": self._validate_stocks,
            "PRODUCTS": self._validate_products,
            "CUSTOMERS": self._validate_customers,
        }
        validators[import_type](records, references, existing)

        invalid_rows = [row for row in records if row["_errors"]]
        duplicate_rows = [
            row for row in records if row["_duplicate"] and not row["_errors"]
        ]
        valid_rows = [
            row
            for row in records
            if not row["_errors"] and not row["_duplicate"]
        ]
        preview = pd.DataFrame(
            [
                {
                    "Ligne": row["_row_number"],
                    "Statut": (
                        "Doublon"
                        if row["_duplicate"]
                        else "Invalide"
                        if row["_errors"]
                        else "Valide"
                    ),
                    "Erreurs": " · ".join(row["_errors"]),
                    **{
                        key: value
                        for key, value in row.items()
                        if not key.startswith("_")
                    },
                }
                for row in records[:100]
            ]
        )
        return {
            "file_name": file_name,
            "file_type": "XLSX",
            "file_hash": hashlib.sha256(content).hexdigest(),
            "import_type": import_type,
            "total_rows": len(records),
            "valid_rows": valid_rows,
            "invalid_rows": invalid_rows,
            "duplicate_rows": duplicate_rows,
            "preview": preview,
            "already_imported": self.repository.file_was_imported(
                hashlib.sha256(content).hexdigest(),
                import_type,
            ),
        }

    def execute_import(
        self,
        analysis: dict,
        import_valid_only: bool = False,
    ) -> dict:
        if analysis["already_imported"]:
            raise ValueError("Ce fichier a déjà été importé avec succès.")
        if analysis["invalid_rows"] and not import_valid_only:
            raise ValueError(
                "Le fichier contient des lignes invalides. Corrigez-le ou "
                "autorisez l'import des seules lignes valides."
            )
        if not analysis["valid_rows"]:
            raise ValueError("Aucune ligne valide et nouvelle à importer.")

        try:
            batch_id, batch_number = self.repository.create_batch(
                file_name=analysis["file_name"],
                file_type=analysis["file_type"],
                import_type=analysis["import_type"],
                file_hash=analysis["file_hash"],
                total_rows=analysis["total_rows"],
                valid_rows=len(analysis["valid_rows"]),
                invalid_rows=len(analysis["invalid_rows"]),
                duplicate_rows=len(analysis["duplicate_rows"]),
            )
            self.repository.save_errors(batch_id, analysis["invalid_rows"])
            imported_rows = self.repository.import_rows(
                analysis["import_type"],
                analysis["valid_rows"],
                batch_id,
            )
            self.repository.complete_batch(
                batch_id,
                imported_rows,
                len(analysis["invalid_rows"]),
            )
            self.db.commit()
        except Exception:
            self.db.rollback()
            raise

        return {
            "batch_id": batch_id,
            "batch_number": batch_number,
            "imported_rows": imported_rows,
            "invalid_rows": len(analysis["invalid_rows"]),
            "duplicate_rows": len(analysis["duplicate_rows"]),
        }

    def get_history(self) -> list[dict]:
        return self.repository.get_history()

    def get_batch_errors(self, batch_id: str) -> list[dict]:
        return self.repository.get_batch_errors(batch_id)

    @staticmethod
    def _read_file(file_name: str, content: bytes) -> pd.DataFrame:
        lower_name = file_name.lower()
        if not lower_name.endswith(".xlsx"):
            raise ValueError("Format non pris en charge. Utilisez un fichier XLSX.")
        excel_file = pd.ExcelFile(io.BytesIO(content))
        sheet_name = "Données" if "Données" in excel_file.sheet_names else 0
        return pd.read_excel(excel_file, sheet_name=sheet_name)

    @staticmethod
    def _build_instructions_sheet(
        sheet,
        import_type: str,
        definition: dict,
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet["A1"] = f"Modèle d'import — {definition['label']}"
        sheet["A1"].font = Font(size=18, bold=True, color="FFFFFF")
        sheet["A1"].fill = PatternFill("solid", fgColor="0F766E")
        sheet.merge_cells("A1:F1")
        sheet["A3"] = "Mode d'emploi"
        sheet["A3"].font = Font(size=13, bold=True, color="0F766E")
        instructions = [
            "1. Ouvrez la feuille Données.",
            "2. Conservez exactement les noms des colonnes.",
            "3. Utilisez les listes déroulantes lorsqu'elles sont proposées.",
            "4. Saisissez une ligne par enregistrement, sans ligne vide intermédiaire.",
            "5. Supprimez la ligne d'exemple si elle ne doit pas être importée.",
            "6. Enregistrez au format .xlsx puis chargez le fichier dans l'application.",
        ]
        for row_index, instruction in enumerate(instructions, start=4):
            sheet.cell(row_index, 1, instruction)
        sheet["A12"] = "Colonnes obligatoires"
        sheet["A12"].font = Font(bold=True, color="FFFFFF")
        sheet["A12"].fill = PatternFill("solid", fgColor="DC2626")
        sheet["A13"] = ", ".join(definition["required"])
        sheet["A15"] = "Colonnes facultatives"
        sheet["A15"].font = Font(bold=True, color="FFFFFF")
        sheet["A15"].fill = PatternFill("solid", fgColor="2563EB")
        sheet["A16"] = ", ".join(definition["optional"])
        sheet["A18"] = "Champs générés automatiquement"
        sheet["A18"].font = Font(bold=True, color="FFFFFF")
        sheet["A18"].fill = PatternFill("solid", fgColor="0F766E")
        sheet["A19"] = ", ".join(definition["automatic"])
        sheet["A21"] = (
            "Les cellules jaunes sont à saisir. Les en-têtes rouges sont "
            "obligatoires et les bleus sont facultatifs."
        )
        sheet.column_dimensions["A"].width = 105
        for row in range(3, 22):
            sheet.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        sheet.row_dimensions[13].height = 42
        sheet.row_dimensions[16].height = 54
        sheet.row_dimensions[19].height = 42
        sheet.row_dimensions[21].height = 36

    @staticmethod
    def _build_references_sheet(
        workbook,
        sheet,
        import_type: str,
        references: dict,
    ) -> None:
        list_sources = EXCEL_LISTS[import_type]
        reference_column = 1
        for field_name, source in list_sources.items():
            if isinstance(source, str):
                values = sorted(references[source])
            else:
                values = list(source)
            sheet.cell(1, reference_column, field_name)
            for row_index, value in enumerate(values, start=2):
                sheet.cell(row_index, reference_column, value)
            safe_name = f"Liste_{field_name}"
            column_letter = get_column_letter(reference_column)
            reference = (
                f"'Références'!${column_letter}$2:"
                f"${column_letter}${max(len(values) + 1, 2)}"
            )
            from openpyxl.workbook.defined_name import DefinedName

            workbook.defined_names.add(
                DefinedName(safe_name, attr_text=reference)
            )
            reference_column += 1

    @staticmethod
    def _build_data_sheet(
        sheet,
        import_type: str,
        definition: dict,
        columns: list[str],
        example: dict,
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1001"

        required = set(definition["required"])
        for column_index, column_name in enumerate(columns, start=1):
            cell = sheet.cell(1, column_index, column_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                "solid",
                fgColor="DC2626" if column_name in required else "2563EB",
            )
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.comment = Comment(
                "Champ obligatoire" if column_name in required else "Champ facultatif",
                "Sales Predictive Analytics",
            )
            example_value = example.get(column_name)
            if example_value is not None and column_name in DATE_COLUMNS:
                example_value = pd.to_datetime(example_value).date()
            elif example_value is not None and column_name in TIME_COLUMNS:
                example_value = pd.to_datetime(str(example_value)).time()
            sheet.cell(2, column_index, example_value)
            sheet.cell(2, column_index).fill = PatternFill(
                "solid", fgColor="FEF3C7"
            )
            width = min(max(len(column_name) + 3, 15), 28)
            sheet.column_dimensions[get_column_letter(column_index)].width = width

            data_range = f"{get_column_letter(column_index)}2:{get_column_letter(column_index)}1001"
            if column_name in EXCEL_LISTS[import_type]:
                validation = DataValidation(
                    type="list",
                    formula1=f"=Liste_{column_name}",
                    allow_blank=column_name not in required,
                )
                validation.errorTitle = "Valeur non autorisée"
                validation.error = "Choisissez une valeur dans la liste."
                validation.promptTitle = column_name
                validation.prompt = "Sélectionnez une valeur dans la liste."
                validation.showErrorMessage = True
                validation.showInputMessage = True
                sheet.add_data_validation(validation)
                validation.add(data_range)
            elif column_name in DATE_COLUMNS:
                validation = DataValidation(
                    type="date",
                    operator="between",
                    formula1="DATE(2020,1,1)",
                    formula2="DATE(2100,12,31)",
                    allow_blank=column_name not in required,
                )
                sheet.add_data_validation(validation)
                validation.add(data_range)
                for row in range(2, 1002):
                    sheet.cell(row, column_index).number_format = "yyyy-mm-dd"
            elif column_name in TIME_COLUMNS:
                for row in range(2, 1002):
                    sheet.cell(row, column_index).number_format = "hh:mm"
            elif column_name in INTEGER_COLUMNS:
                validation = DataValidation(
                    type="whole",
                    operator="greaterThan",
                    formula1="0",
                    allow_blank=column_name not in required,
                )
                sheet.add_data_validation(validation)
                validation.add(data_range)
            elif column_name in POSITIVE_COLUMNS:
                validation = DataValidation(
                    type="decimal",
                    operator="greaterThan",
                    formula1="0",
                    allow_blank=column_name not in required,
                )
                sheet.add_data_validation(validation)
                validation.add(data_range)
            elif column_name in NON_NEGATIVE_COLUMNS:
                validation = DataValidation(
                    type="decimal",
                    operator="greaterThanOrEqual",
                    formula1="0",
                    allow_blank=column_name not in required,
                )
                sheet.add_data_validation(validation)
                validation.add(data_range)

        last_column = get_column_letter(len(columns))
        sheet.conditional_formatting.add(
            f"A2:{last_column}1001",
            FormulaRule(
                formula=["COUNTA($A2:$ZZ2)>0"],
                fill=PatternFill("solid", fgColor="FFF7ED"),
            ),
        )
        sheet.row_dimensions[1].height = 30

    @staticmethod
    def _normalize_column(value) -> str:
        normalized = unicodedata.normalize("NFKD", str(value).strip().lower())
        normalized = "".join(
            char for char in normalized if not unicodedata.combining(char)
        )
        return re.sub(r"[^a-z0-9]+", "_", normalized).strip("_")

    @staticmethod
    def _clean_value(value):
        if pd.isna(value):
            return None
        if isinstance(value, str):
            value = value.strip()
            return value or None
        return value

    @staticmethod
    def _number(row: dict, column: str, default=None, integer=False):
        value = row.get(column)
        if value is None:
            return default
        try:
            number = float(str(value).replace(" ", "").replace(",", "."))
            return int(number) if integer else number
        except (TypeError, ValueError):
            row["_errors"].append(f"{column} doit être numérique.")
            return default

    @staticmethod
    def _date(row: dict, column: str):
        try:
            value = row.get(column)
            if isinstance(value, str) and re.fullmatch(
                r"\d{4}-\d{2}-\d{2}", value
            ):
                return pd.to_datetime(value, format="%Y-%m-%d").date()
            return pd.to_datetime(value, dayfirst=True).date()
        except (TypeError, ValueError):
            row["_errors"].append(f"{column} doit contenir une date valide.")
            return None

    @staticmethod
    def _boolean(value, default=False) -> bool:
        if value is None:
            return default
        if isinstance(value, bool):
            return value
        return str(value).strip().lower() in {
            "1", "true", "vrai", "yes", "oui", "o", "x",
        }

    def _validate_sales(self, rows, references, existing):
        seen_lines = set()
        invalid_sales = set()
        sale_signatures = {}
        for row in rows:
            sale_reference = str(row.get("sale_reference") or "").strip()
            product_code = str(row.get("product_code") or "").strip()
            row["sale_reference"] = sale_reference
            row["product_code"] = product_code
            if not sale_reference:
                row["_errors"].append("sale_reference est obligatoire.")
            elif len(sale_reference) > 100:
                row["_errors"].append(
                    "sale_reference ne peut pas dépasser 100 caractères."
                )
            if sale_reference in existing:
                row["_duplicate"] = True
            line_key = (sale_reference, product_code)
            if line_key in seen_lines:
                row["_errors"].append(
                    "Produit présent plusieurs fois dans la même vente."
                )
            seen_lines.add(line_key)

            product = references["products"].get(product_code)
            if product is None:
                row["_errors"].append(f"Produit inconnu : {product_code}.")
            customer_code = str(row.get("customer_code") or "").strip()
            customer = references["customers"].get(customer_code)
            if customer_code and customer is None:
                row["_errors"].append(f"Client inconnu : {customer_code}.")

            row["sale_date"] = self._date(row, "sale_date")
            if row["sale_date"] and row["sale_date"] > date.today():
                row["_errors"].append(
                    "sale_date ne peut pas être dans le futur."
                )
            raw_time = row.get("sale_time")
            if raw_time is None:
                row["sale_time"] = None
            elif isinstance(raw_time, time):
                row["sale_time"] = raw_time
            else:
                try:
                    row["sale_time"] = pd.to_datetime(str(raw_time)).time()
                except ValueError:
                    row["_errors"].append("sale_time doit être une heure valide.")
                    row["sale_time"] = None

            row["quantity_packages"] = self._number(
                row, "quantity_packages"
            )
            row["unit_price"] = self._number(row, "unit_price")
            row["discount_amount"] = self._number(
                row, "discount_amount", default=0.0
            )
            for column in ("quantity_packages", "unit_price"):
                if row[column] is not None and row[column] <= 0:
                    row["_errors"].append(f"{column} doit être supérieur à 0.")
            line_subtotal = (row["quantity_packages"] or 0) * (
                row["unit_price"] or 0
            )
            if row["discount_amount"] < 0 or row["discount_amount"] > line_subtotal:
                row["_errors"].append("Remise invalide pour cette ligne.")

            payment_method = row.get("payment_method") or "CASH"
            payment_status = row.get("payment_status") or "PAID"
            signature = (
                row["sale_date"],
                customer_code,
                payment_method,
                payment_status,
            )
            previous_signature = sale_signatures.setdefault(
                sale_reference,
                signature,
            )
            if signature != previous_signature:
                row["_errors"].append(
                    "Les lignes d'une même vente doivent avoir la même "
                    "date, le même client et le même paiement."
                )

            row.update(
                {
                    "product_id": product["id"] if product else None,
                    "product_name": product["name"] if product else None,
                    "minimum_stock": (
                        float(product["minimum_stock"]) if product else 0
                    ),
                    "customer_id": customer["id"] if customer else None,
                    "units_per_package": (
                        int(product["units_per_package"]) if product else 1
                    ),
                    "unit_cost": float(product["purchase_price"]) if product else 0,
                    "line_subtotal": line_subtotal,
                    "salesperson_name": row.get("salesperson_name"),
                    "payment_method": payment_method,
                    "payment_status": payment_status,
                    "promotion_applied": row["discount_amount"] > 0,
                    "external_reference": sale_reference,
                    "notes": row.get("notes"),
                }
            )
            if row["_errors"]:
                invalid_sales.add(sale_reference)

        for row in rows:
            if row["sale_reference"] in invalid_sales and not row["_errors"]:
                row["_errors"].append(
                    "Une autre ligne de cette vente est invalide."
                )

    def _validate_stocks(self, rows, references, existing):
        seen = set()
        for row in rows:
            product_code = str(row.get("product_code") or "").strip()
            product = references["products"].get(product_code)
            if product is None:
                row["_errors"].append(f"Produit inconnu : {product_code}.")
            row["stock_date"] = self._date(row, "stock_date")
            if row["stock_date"] and row["stock_date"] > date.today():
                row["_errors"].append(
                    "stock_date ne peut pas être dans le futur."
                )
            key = (row["stock_date"], product_code)
            if key in existing or key in seen:
                row["_duplicate"] = True
            seen.add(key)
            for column in (
                "opening_stock", "quantity_received", "quantity_sold",
                "quantity_damaged", "other_entries", "other_outputs",
                "closing_stock",
            ):
                default = None if column in {"opening_stock", "closing_stock"} else 0.0
                row[column] = self._number(row, column, default=default)
                if row[column] is not None and row[column] < 0:
                    row["_errors"].append(f"{column} ne peut pas être négatif.")
            if row["opening_stock"] is not None and row["closing_stock"] is not None:
                expected = (
                    row["opening_stock"] + row["quantity_received"]
                    + row["other_entries"] - row["quantity_sold"]
                    - row["quantity_damaged"] - row["other_outputs"]
                )
                if abs(expected - row["closing_stock"]) > 0.01:
                    row["_errors"].append(
                        f"Stock de clôture incohérent (attendu : {expected:.2f})."
                    )
            row.update(
                {
                    "product_id": product["id"] if product else None,
                    "minimum_stock": (
                        float(product["minimum_stock"]) if product else 0
                    ),
                    "stockout_flag": (row["closing_stock"] or 0) <= 0,
                }
            )

    def _validate_products(self, rows, references, existing):
        seen = set()
        for row in rows:
            category_code = str(row.get("category_code") or "").strip()
            category = references["categories"].get(category_code)
            if not row.get("name"):
                row["_errors"].append("name est obligatoire.")
            if category is None:
                row["_errors"].append(f"Catégorie inconnue : {category_code}.")
            row["units_per_package"] = self._number(
                row, "units_per_package", integer=True
            )
            row["selling_price"] = self._number(row, "selling_price")
            for column in (
                "volume_value", "purchase_price", "minimum_stock",
                "reorder_quantity",
            ):
                row[column] = self._number(row, column, default=0.0)
            if not row.get("package_type"):
                row["_errors"].append("package_type est obligatoire.")
            if not row["units_per_package"] or row["units_per_package"] <= 0:
                row["_errors"].append("units_per_package doit être supérieur à 0.")
            if not row["selling_price"] or row["selling_price"] <= 0:
                row["_errors"].append("selling_price doit être supérieur à 0.")
            identity = (
                self._normalized_text(row.get("name")),
                self._normalized_text(row.get("brand")),
                row["volume_value"],
                self._normalized_text(row.get("volume_unit")),
                self._normalized_text(row.get("package_type")),
            )
            if identity in existing or identity in seen:
                row["_duplicate"] = True
            seen.add(identity)
            row.update(
                {
                    "category_id": category["id"] if category else None,
                    "brand": row.get("brand"),
                    "volume_unit": row.get("volume_unit"),
                }
            )

    def _validate_customers(self, rows, references, existing):
        seen_phones = set()
        seen_identities = set()
        for row in rows:
            type_code = str(row.get("customer_type_code") or "").strip()
            customer_type = references["customer_types"].get(type_code)
            if not row.get("name"):
                row["_errors"].append("name est obligatoire.")
            if customer_type is None:
                row["_errors"].append(f"Type de client inconnu : {type_code}.")
            phone = self._normalized_phone(row.get("phone"))
            if phone and len(phone) < 8:
                row["_errors"].append(
                    "Le numéro de téléphone doit contenir au moins 8 chiffres."
                )
            identity = (
                self._normalized_text(row.get("name")),
                self._normalized_text(row.get("district")),
                self._normalized_text(row.get("city") or "Bamako"),
            )
            if phone and (
                phone in existing["phones"] or phone in seen_phones
            ):
                row["_duplicate"] = True
            if (
                identity in existing["identities"]
                or identity in seen_identities
            ):
                row["_duplicate"] = True
            if phone:
                seen_phones.add(phone)
            seen_identities.add(identity)
            row.update(
                {
                    "customer_type_id": (
                        customer_type["id"] if customer_type else None
                    ),
                    "phone": row.get("phone"),
                    "zone": row.get("zone"),
                    "district": row.get("district"),
                    "city": row.get("city") or "Bamako",
                }
            )

    @staticmethod
    def _normalized_text(value) -> str:
        return " ".join(str(value or "").strip().lower().split())

    @staticmethod
    def _normalized_phone(value) -> str:
        return re.sub(r"[^0-9]", "", str(value or ""))
