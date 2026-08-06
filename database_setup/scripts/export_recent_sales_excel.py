"""Exporte les ventes récentes d'un dépôt dans le modèle Excel NexaStock."""

import argparse
import io
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import text

PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.database.session import session_for_company
from app.services.data_import_service import DataImportService


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--company-id", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--reference-prefix", default="NEXA-BERTHE-%")
    return parser.parse_args()


def load_rows(db, reference_prefix):
    return list(db.execute(text("""
        SELECT
            s.external_reference AS sale_reference,
            s.sale_date,
            s.sale_time,
            COALESCE(c.code, '') AS customer_code,
            p.code AS product_code,
            si.quantity_packages,
            si.unit_price,
            si.discount_amount,
            COALESCE(s.payment_method, 'CASH') AS payment_method,
            COALESCE(s.payment_status, 'PAID') AS payment_status,
            COALESCE(s.salesperson_name, '') AS salesperson_name,
            COALESCE(s.notes, '') AS notes
        FROM sales s
        JOIN sale_items si
          ON si.company_id = s.company_id AND si.sale_id = s.id
        JOIN products p
          ON p.company_id = si.company_id AND p.id = si.product_id
        LEFT JOIN customers c
          ON c.company_id = s.company_id AND c.id = s.customer_id
        WHERE s.external_reference LIKE :reference_prefix
          AND s.deleted_at IS NULL
        ORDER BY s.sale_date, s.sale_time, s.external_reference, p.code
    """), {"reference_prefix": reference_prefix}).mappings())


def copy_row_style(sheet, source_row, target_row, column_count):
    for column in range(1, column_count + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)


def build_workbook(company_id, output_path, reference_prefix):
    with session_for_company(company_id) as db:
        service = DataImportService(db)
        template = service.get_template("SALES", "XLSX")
        rows = load_rows(db, reference_prefix)

    if not rows:
        raise RuntimeError("Aucune vente récente ne correspond au préfixe demandé.")

    workbook = load_workbook(io.BytesIO(template))
    sheet = workbook["Données"]
    columns = [cell.value for cell in sheet[1]]
    column_indexes = {name: index for index, name in enumerate(columns, start=1)}

    for row_index, row in enumerate(rows, start=2):
        copy_row_style(sheet, 2, row_index, len(columns))
        for name in columns:
            value = row.get(name)
            cell = sheet.cell(row_index, column_indexes[name], value)
            if name == "sale_date":
                cell.number_format = "yyyy-mm-dd"
            elif name == "sale_time":
                cell.number_format = "hh:mm"
            elif name in {"quantity_packages", "unit_price", "discount_amount"}:
                cell.number_format = "#,##0.00"

    last_row = len(rows) + 1
    last_column_letter = sheet.cell(1, len(columns)).column_letter
    preferred_widths = {
        "sale_reference": 34,
        "sale_date": 14,
        "product_code": 18,
        "quantity_packages": 20,
        "unit_price": 16,
        "sale_time": 12,
        "customer_code": 18,
        "salesperson_name": 30,
        "payment_method": 20,
        "payment_status": 18,
        "discount_amount": 18,
        "notes": 42,
    }
    for name, width in preferred_widths.items():
        sheet.column_dimensions[sheet.cell(1, column_indexes[name]).column_letter].width = width
    for row_index in range(2, last_row + 1):
        sheet.row_dimensions[row_index].height = 20
        for column_index in range(1, len(columns) + 1):
            sheet.cell(row_index, column_index).alignment = Alignment(
                vertical="center",
                wrap_text=False,
            )
    sheet.auto_filter.ref = f"A1:{last_column_letter}{last_row}"
    sheet.freeze_panes = "A2"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:1"
    sheet.print_area = f"A1:{last_column_letter}{last_row}"

    instructions = workbook["Instructions"]
    instructions["A23"] = "Contenu du fichier"
    instructions["A23"].font = Font(bold=True, color="FFFFFF")
    instructions["A23"].fill = PatternFill("solid", fgColor="0F766E")
    instructions["A24"] = (
        f"{len(rows)} lignes de vente récentes exportées. Ce fichier reflète "
        "les données déjà présentes dans la base locale Berthe KLB : ne le "
        "réimportez pas dans cette même base, où elles seront reconnues comme doublons."
    )
    instructions["A24"].alignment = Alignment(wrap_text=True, vertical="top")
    instructions.row_dimensions[24].height = 58

    output_path = Path(output_path).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.active = workbook.sheetnames.index("Données")
    workbook.save(output_path)
    return output_path, len(rows), sheet.max_column, last_row


if __name__ == "__main__":
    args = parse_args()
    path, row_count, column_count, last_row = build_workbook(
        args.company_id,
        args.output,
        args.reference_prefix,
    )
    print(
        f"Classeur créé : {path} · {row_count} lignes · "
        f"{column_count} colonnes · dernière ligne {last_row}"
    )
